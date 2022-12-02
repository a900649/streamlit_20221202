import pandas as pd
import streamlit as st
import plotly.express as px
from astropy import units as u
from astroquery.nasa_exoplanet_archive import NasaExoplanetArchive
import time
import numpy as np


import electricity_statistics
st.title('Home')
st.markdown('---')

page_names_to_funcs = {
    "用電戶數": electricity_statistics.show_用電戶數統計數據,
    "使用電量": electricity_statistics.show_使用電量,
}

uploaded_file = st.file_uploader(label="請上傳一個檔案",type=["xlsx","xls"])
if uploaded_file is not None:
    df = pd.read_excel(uploaded_file)

#########################################################
# import xlsxwriter
# from io import BytesIO
#
# a = pd.DataFrame([[1,2],[3,4,]],columns=["A","B"])
# output = BytesIO()
#
# workbook = xlsxwriter.Workbook(output, {'in_memory': True})
# worksheet = workbook.add_worksheet("s1")
#
# # 填Columns
# for col in range(0,len(a.columns)):
#
#     worksheet.write(0, col, a.columns[col])
#
# # 填數據
# for ind in range(0,len(a.index)):
#     for col in range(0,len(a.columns)):
#         worksheet.write(ind+1,col, a.iloc[ind,col])
#
# workbook.close()
#
# st.download_button(
#     label="Download Excel workbook",
#     data=output.getvalue(),
#     file_name="workbook.xlsx",
#     mime="application/vnd.ms-excel"
# )
#########################################################

#########################################################

# from io import BytesIO
#
# towrite = BytesIO()
# a = pd.DataFrame({'A':[1,2,3],'B':[4,5,6]})
# a.to_excel(towrite,sheet_name="aaa")
#
# towrite.seek(0)  # reset pointer
#
# st.download_button(
#     label="Download Excel workbook",
#     data=towrite.getvalue(),
#     file_name="workbook.xlsx",
#     mime="application/vnd.ms-excel"
# )

#########################################################

import openpyxl as op
from openpyxl.styles import Alignment ,Font ,NamedStyle,PatternFill,Side,Border
from openpyxl.utils import get_column_letter
from io import BytesIO

wb = op.Workbook()
ws = wb.create_sheet(title='Sheet2',index=0)
ws.cell(1,1).value = "A"
ws.cell(1,2).value = "b"

ws.cell(2,1).value = "1"
ws.cell(2,2).value = "2"


towrite = BytesIO()
wb.save(towrite)
towrite.seek(0)  # reset pointer

st.download_button(
    label="Download Excel workbook",
    data=towrite.getvalue(),
    file_name="workbook.xlsx",
    mime="application/vnd.ms-excel"
)

#########################################################




demo_name = st.sidebar.selectbox("選擇頁面", page_names_to_funcs.keys())
page_names_to_funcs[demo_name]()



